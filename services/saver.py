import openpyxl


def save_to_csv(products: list, fields: list, path: str='./', filename: str='table') -> None:
    f = open(f'{path}{filename}.csv', 'w')
    f.write(';'.join(fields) + '\n')
    for product in products:
        f.write(';'.join(product.values()) + '\n')


def save_to_exel(products: list, fields: list, path: str='./', filename: str='table'):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Задание'

    for i, value in enumerate(fields):
        sheet.cell(row=1, column=i+1).value = value

    for row_num, p in enumerate(products):
        for row_col, value in enumerate(p.values()):
            sheet.cell(row=2+row_num, column=row_col+1).value = value

    wb.save(f'{path}{filename}.xlsx')

